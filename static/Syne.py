import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import win32com
import win32com.client as win32
import matplotlib.pyplot as plt
import subprocess
import os
import time
import platform
import paramiko as paramiko
import argparse
import shutil
import requests
from datetime import datetime

import os
import time
from django.core.files import File


def ssh_execute_script(
    client_IP,
    client_name,
    client_password,
    remote_base_path,
    local_folder_path,
    local_bat_file,
):
    # SSH 세션 시작
    ssh = paramiko.SSHClient()

    # 호스트 메소드 정보 입력
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # SSH를 하기 위한 원격 컴퓨터 정보
    ssh.connect(client_IP, username=client_name, password=client_password)

    # SSH를 통해 점검 프로그램이 들어갈 경로를 생성
    create_folder_command = "mkdir C:\\informationSS3"
    stdizn, stdout, stderr = ssh.exec_command(create_folder_command)

    # SFTP 세션 시작
    sftp = ssh.open_sftp()

    # SFTP를 통해 로컬에 있던 점검프로그램을 원격 컴퓨터의 특정 경로에다 삽입
    sftp.put(local_bat_file, "C:\\informationSS3\\Windo.bat")

    # 만들어둔 경로를 통해, 그 경로의 점검 프로그램을 원격에서 실행
    command = "cd /d C:\\informationSS3 && Windo.bat"
    stdin, stdout, stderr = ssh.exec_command(command)

    # 원격 점검에서 생성되는 SeScore3.txt 가 있는지 확인. 없으면 생길 때 까지 5초동안 대기
    remote_file_to_check = "C:\\informationSS3\\W1~82\\Score\\SeScore3.txt"
    while True:
        try:
            sftp.stat(remote_file_to_check)
            break
        except FileNotFoundError:
            print("점검이 진행중입니다...")
            time.sleep(5)

    # 원격 점검프로그램에서 생성되는 디렉토리 이름들을 구조체로 선언
    subdirectories = ["action", "bad", "good", "log", "Score"]

    # 위에서 선언한 구조체 하나하나를 subdir로 반복해서 아래 함수로 사용
    # remote_base_path = 'C:\\informationSS3\\W1~82' 가 큰 경로라 하면
    # os.path.join(remote_base_path, subdir) 는
    # C:\\informationSS3\\W1~82\\action ...  C:\\informationSS3\\W1~82\\bad..
    # 등으로 반환되는 방식

    for subdir in subdirectories:
        remote_folder_path = os.path.join(remote_base_path, subdir)
        local_subdir_path = os.path.join(local_folder_path, subdir)

        # 로컬에 넣을 경로가 없다면 경로 생성
        if not os.path.exists(local_subdir_path):
            os.makedirs(local_subdir_path)

        # 원격 컴퓨터의 생성파일들을 리스트화 시킴.
        # 만약 오류가 생기면, 생성파일의 이름을 부르고 오류가 났다고 출력
        # 출력이 끝나면 계속 하던거 계속하도록함

        try:
            remote_files = sftp.listdir(remote_folder_path)
        except IOError as e:
            print(f"Error: {e}. Skipping {remote_folder_path}.")
            continue

        # 원격의 각 파일들을 아까 리스트화 된 것을 통해
        # 로컬 폴더에다가 복사해서 가져옴

        for remote_file in remote_files:
            remote_file_path = os.path.join(remote_folder_path, remote_file)
            local_file_path = os.path.join(local_subdir_path, remote_file)
            sftp.get(remote_file_path, local_file_path)

    # report.txt 파일은 따로 복사해서 가져옴
    sftp.get(
        os.path.join(remote_base_path, "report.txt"),
        os.path.join(local_folder_path, "report.txt"),
    )

    # SFTP 세션 종료
    sftp.close()

    # 원격에서 만든 C:\\informationSS3" 이 폴더를 삭제
    # 그럼 원격에서 아무런 흔적이 남지 않을 것임
    # 먼저 지금 실행경로가 C드라이브안의 폴더이기 떄문에, 경로를 벗어나야함.
    # 그렇지 않으면 사용중일라고 안될 우려가 있음

    command1 = "cd C:\\"
    stdin, stdout, stderr = ssh.exec_command(command1)
    delete_command1 = "rd /s /q C:\\informationSS3"
    stdin, stdout, stderr = ssh.exec_command(delete_command1)

    # SSH 세션 종
    ssh.close()


# 함수로 정의한 SSH / SFTP 사용
# 그리고 그에 대한 정보

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("client_IP", help="원격 컴퓨터 IP 정보")
    parser.add_argument("client_name", help="관리자 아이디")
    parser.add_argument("client_password", help="관리자 비밀번호")
    args = parser.parse_args()

    # 현재 시간과 날짜.
    now = datetime.now()
    # 폴더명으로 사용할 형식을 지정.. EX): "YYYY-MM-DD_HH-MM-SS"
    folder_name0 = now.strftime("%Y-%m-%d_%H-%M-%S")
    # 새로운 폴더를 만들고자 하는 경로
    Media_path = os.environ.get("FOLDER_PATH")
    # Media_path = (
    #     r"C:\Users\VinoMatinee\Desktop\projectCTF-main\media\syne" + folder_name0
    # )
    # 폴더가 존재하지 않을 경우에만 폴더를 생성.
    try:
        if not os.path.exists(Media_path):
            os.makedirs(Media_path)
        if not os.path.exists(Media_path + "\\img"):
            os.makedirs(Media_path + "\\img")
        if not os.path.exists(Media_path + "\\txt"):
            os.makedirs(Media_path + "\\txt")
        if not os.path.exists(Media_path + "\\Solution"):
            os.makedirs(Media_path + "\\Solution")
        print("폴더 생성 완료...")
    except:
        print("사용자 폴더 생성 실패")
        print("조치를 취하십시오.")

    client_IP = args.client_IP
    client_name = args.client_name
    client_password = args.client_password
    remote_base_path = "C:\\informationSS3\\W1~82"  # 원격 컴퓨터 경로
    local_folder_path = Media_path + "\\txt"  # 원격 컴퓨터에서 복사한 파일을 옮길 로컬 경로
    local_bat_file = "static\\Windo_main.bat"  # 원격 컴퓨터에 넣을 점검 파일 경로

    ssh_execute_script(
        client_IP,
        client_name,
        client_password,
        remote_base_path,
        local_folder_path,
        local_bat_file,
    )


# 원격에서 가져온 파일들 중에 점수가 적혀있는 텍스트 정보
# 각각 오픈에서 읽어서 객체선언

with open(local_folder_path + r"\\Score\\AScore.txt", "r", encoding="ANSI") as f:
    Ascore = int(f.read())
AscorePer = round((Ascore / 147) * 100, 2)

with open(local_folder_path + r"\\Score\\SScore.txt", "r", encoding="ANSI") as f:
    Sscore = int(f.read())
SscorePer = round((Sscore / 348) * 100, 2)

with open(local_folder_path + r"\\Score\\PScore.txt", "r", encoding="ANSI") as f:
    Pscore = int(f.read())
PscorePer = round((Pscore / 9) * 100, 2)

with open(local_folder_path + r"\\Score\\LScore.txt", "r", encoding="ANSI") as f:
    Lscore = int(f.read())
LscorePer = round((Lscore / 27) * 100, 2)

with open(local_folder_path + r"\\Score\\SeScore.txt", "r", encoding="ANSI") as f:
    Sescore = int(f.read())
SescorePer = round((Sescore / 168) * 100, 2)


# 사용할 색상 선언
colors = ["#66b3ff", "#ff9999"]

## 계정 항목 그래프 ##

# 도넛 그래프 그리기
fig, ax = plt.subplots()
ax.pie([AscorePer, 100 - AscorePer], colors=colors, startangle=90)
# 중앙에 원 그리기
centre_circle = plt.Circle((0, 0), 0.70, fc="white")
fig.gca().add_artist(centre_circle)
# 도넛 구멍에 값 표시
plt.text(0, 0, f"{AscorePer}%", ha="center", fontsize=25)
# 원형 유지를 위해 'equal'로 설정
plt.axis("equal")
# 그래프를 저장
plt.savefig(Media_path + r"\\img\\Account_Chart.png")  # 이미지 저장

## 서비스 항목 그래프 ##
fig2, ax2 = plt.subplots()
ax2.pie([SscorePer, 100 - SscorePer], colors=colors, startangle=90)
centre_circle2 = plt.Circle((0, 0), 0.70, fc="white")
fig2.gca().add_artist(centre_circle2)
plt.text(0, 0, f"{SscorePer}%", ha="center", fontsize=25)
plt.axis("equal")
plt.savefig(Media_path + r"\\img\\Service_Chart.png")

## 패치 항목 그래프 ##
fig3, ax3 = plt.subplots()
ax3.pie([PscorePer, 100 - PscorePer], colors=colors, startangle=90)
centre_circle3 = plt.Circle((0, 0), 0.70, fc="white")
fig3.gca().add_artist(centre_circle3)
plt.text(0, 0, f"{PscorePer}%", ha="center", fontsize=25)
plt.axis("equal")
plt.savefig(Media_path + r"\\img\\Patch_Chart.png")

## 로그 항목 그래프 ##
fig4, ax4 = plt.subplots()
ax4.pie([LscorePer, 100 - LscorePer], colors=colors, startangle=90)
centre_circle4 = plt.Circle((0, 0), 0.70, fc="white")
fig4.gca().add_artist(centre_circle4)
plt.text(0, 0, f"{LscorePer}%", ha="center", fontsize=25)
plt.axis("equal")
plt.savefig(Media_path + r"\\img\\Log_Chart.png")

## 서비스 항목 그래프 ##
fig5, ax5 = plt.subplots()
ax5.pie([SescorePer, 100 - SescorePer], colors=colors, startangle=90)
centre_circle5 = plt.Circle((0, 0), 0.70, fc="white")
fig5.gca().add_artist(centre_circle5)
plt.text(0, 0, f"{SescorePer}%", ha="center", fontsize=25)
plt.axis("equal")
plt.savefig(Media_path + r"\\img\\Secure_Chart.png")

# Workbook이란 이름의 엑셀파일을 만듬
wb = openpyxl.Workbook()

# 파일 이름을 재설정
new_filename = Media_path + r"\\Solution\\주요통신기반.xlsx"

# 시트 제목을 정함
wb.active.title = "Report"

# 첫번째 시트를 변수로 쓰기 편하게 지정
w1 = wb["Report"]

r = 21

file = open(local_folder_path + r"\\report.txt", encoding="ANSI")
while True:
    line = file.readline()
    if not line:
        break
    if not line.strip():  # 공백 문자열인 경우 continue를 호출하여 다음 루프로 이동
        continue
    w1.cell(row=r, column=1).value = line.strip()
    r = r + 1
file.close()

# Report 시트의 형식을 미리 지정
w1.merge_cells("A1:C1")  # Account Score 부분 병합
w1.cell(row=1, column=1).value = "Account Score"
w1.cell(row=1, column=1).alignment = Alignment(horizontal="center")  # 가운데 정렬

w1.merge_cells("D1:F1")  # Service Score 부분 병합
w1.cell(row=1, column=4).value = "Service Score"
w1.cell(row=1, column=4).alignment = Alignment(horizontal="center")  # 가운데 정렬

w1.merge_cells("G1:I1")  # Patch Score 부분 병합
w1.cell(row=1, column=7).value = "Patch Score"
w1.cell(row=1, column=7).alignment = Alignment(horizontal="center")  # 가운데 정렬

w1.merge_cells("B9:C9")  # Log Score 부분 병합
w1.cell(row=9, column=2).value = "Log Score"
w1.cell(row=9, column=2).alignment = Alignment(horizontal="center")  # 가운데 정렬

w1.merge_cells("F9:G9")  # Secure Score 부분 병합
w1.cell(row=9, column=6).value = "Secure Score"
w1.cell(row=9, column=6).alignment = Alignment(horizontal="center")  # 가운데 정렬


##계정 항목 이미지##

# 이미지 파일을 엑셀에 삽입하고 크기 설정
img = Image(Media_path + r"\\img\\Account_Chart.png")
img.width = 3 * 70  # 가로 크기 설정 (4 셀 = 4*64 픽셀)
img.height = 9 * 15  # 세로 크기 설정 (9 셀 = 9*15 픽셀)
w1.add_image(img, "A2")  # A2 셀에 이미지 삽입

## 서비스 항목 이미지 ##
img2 = Image(Media_path + r"\\img\\Service_Chart.png")
img2.width = 3 * 70
img2.height = 9 * 15
w1.add_image(img2, "D2")
## 패치 항목 이미지 ##
img3 = Image(Media_path + r"\\img\\Patch_Chart.png")
img3.width = 3 * 70
img3.height = 9 * 15
w1.add_image(img3, "G2")
## 로그 항목 이미지 ##
img4 = Image(Media_path + r"\\img\\Log_Chart.png")
img4.width = 4 * 70
img4.height = 14 * 15
w1.add_image(img4, "A10")
## 보안 항목 이미지 ##
img5 = Image(Media_path + r"\\img\\Secure_Chart.png")
img5.width = 4 * 70
img5.height = 14 * 15
w1.add_image(img5, "E10")

"""
# 페이지 여백 조절
ws.page_margins.left = 0.1
ws.page_margins.right = 0.1
ws.page_margins.top = 0.1
ws.page_margins.bottom = 0.1
"""


# 엑셀 변경사항 저장
wb.save(new_filename)

excel = win32com.client.Dispatch("Excel.Application")  # 엑셀 어플리케이션 백그라운드로 실행
wb = excel.Workbooks.open(Media_path + r"\Solution\주요통신기반.xlsx")  # 엑셀 파일을 읽어드려서 객체로 지정
ws_Report = wb.Worksheets("Report")
ws_Report.Select()
pdf_path = Media_path + r"\Solution\주요통신기반.pdf"
wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)  # 지정했던 경로로 pdf 파일 생성
wb.Close(False)  # 엑셀 작업을 종료시키고 객체를 시스템에 반환
excel.Quit()  # 백그라운드로 켜져있는 엑셀을 종료. 이 문구 없으면 백그라운드로 실행되기 때문에 작업관리자 켜서, 엑셀을 직접 꺼야함.

print("작업 종료")
